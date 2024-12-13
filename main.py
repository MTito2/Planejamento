import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
import warnings
import os

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def ler_planilha(caminho_arquivo, nome_aba_especifica):
    print("Analisando dados, aguarde...")
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb.active
    if nome_aba_especifica not in wb.sheetnames:
        raise ValueError(f"A aba '{nome_aba_especifica}' não foi encontrada na planilha.")
    
    sheet = wb[nome_aba_especifica]
    nome_aba = sheet.title
    dados_colunas = {}
    linha_inicial = 1
    linha_final = ws.max_row
    
    for col in range(1, sheet.max_column + 1):
        coluna = []
        for i in range(linha_inicial, linha_final + 1):
            valor = sheet.cell(row=i, column=col).value
            if valor is None:
                valor = 0  # Tratar células vazias como 0
            coluna.append(valor)

        dados_colunas[f'Coluna_{col}'] = coluna
        
    dados_colunas.popitem()
    return nome_aba, dados_colunas

def processar_dados(dados):
    agrupados = {}
    for chave, valores in dados.items():
        coluna_index = int(chave.split("_")[1])
        if coluna_index < 9:
            continue
        
        subtopicos = dados["Coluna_3"][3:]  # A partir da linha 4
        periodo = valores[0].strftime("%d/%m/%Y")  # Formatação de data
        valores_somados = valores[3:]  # A partir da linha 4
        valores_somados = [float(valor) if valor not in (None, '', ' ') else 0 for valor in valores_somados]

        if periodo in agrupados:
            for i, valor in enumerate(valores_somados):
                if i < len(agrupados[periodo]):
                    agrupados[periodo][i] += valor
                else:
                    agrupados[periodo].append(valor)
        else:
            agrupados[periodo] = valores_somados
    
    return subtopicos, agrupados

def unir_subtopicos_valores(subtopicos, agrupados):
    unidos = {}
    for periodo, valores in agrupados.items():
        if len(valores) != len(subtopicos):
            raise ValueError(f"Inconsistência nos tamanhos para o período {periodo}: "
                             f"Valores ({len(valores)}) e Subtópicos ({len(subtopicos)}).")
        unidos[periodo] = list(zip(subtopicos, valores))
    return unidos

def criar_planilha(dados_unidos, nome_arquivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Processados"

    # Obter os períodos (colunas) e sub-tópicos (linhas)
    periodos = list(dados_unidos.keys())
    subtopicos = [item[0] for item in dados_unidos[next(iter(dados_unidos))]]

    # Adicionar bordas
    borda = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Criar o estilo de alinhamento centralizado
    alinhamento = Alignment(horizontal="center", vertical="center")

    # Escrever os períodos na primeira linha (a partir da segunda coluna) e aplicar alinhamento e borda
    for col_idx, periodo in enumerate(periodos, start=2):
        cell = ws.cell(row=1, column=col_idx, value=periodo)
        cell.border = borda  # Adiciona borda
        cell.alignment = alinhamento  # Centraliza o texto

    # Escrever os sub-tópicos na primeira coluna (a partir da segunda linha) e aplicar alinhamento e borda
    for row_idx, subtopico in enumerate(subtopicos, start=2):
        cell = ws.cell(row=row_idx, column=1, value=subtopico)
        cell.border = borda  # Adiciona borda
        cell.alignment = alinhamento  # Centraliza o texto

    # Preencher os valores, adicionar bordas e centralizar
    for col_idx, periodo in enumerate(periodos, start=2):
        valores = dados_unidos[periodo]
        for row_idx, (subtopico, valor) in enumerate(valores, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = borda  # Adiciona borda
            cell.alignment = alinhamento  # Centraliza o texto

    # Ajustar a largura das colunas automaticamente
    for col_idx in range(1, len(periodos) + 2):  # +2 para incluir a coluna de sub-tópicos
        max_length = 0
        column = ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)]
        
        # Determina a maior largura para cada coluna
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        column.width = adjusted_width
    os.system("cls")
    # Salvar o arquivo
    wb.save(nome_arquivo)
    print(f"Planilha salva em: {nome_arquivo}")


# Definindo o caminho e parâmetros
arquivo = input("Informe o nome do arquivo: ")

caminho = f"C:\\Users\\SALUM\\Documents\\curva_financeira\\{arquivo}.xlsx"
nome_aba = 'CRN'

# Ler a planilha
nome_aba, dados = ler_planilha(caminho, nome_aba)

# Processar os dados
subtopicos, agrupados = processar_dados(dados)

# Unir sub-tópicos e valores
dados_unidos = unir_subtopicos_valores(subtopicos, agrupados)

# Definir o nome do arquivo e intervalos a excluir
nome_arquivo = r"C:\\Users\\SALUM\\Documents\\curva_financeira\\Auxiliar de atualização financeira.xlsx"
# Criar a planilha final
criar_planilha(dados_unidos, nome_arquivo)
